'''report_milk.report_milk_xlsx.py'''

import inspect
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from container import get_dependency


class ReportMilkXlsx:
    def __init__(self):
        print(f"ReportMilkXlsx instantiated by: {inspect.stack()[1].filename}")
        self.MA = None
        self.WG = None
        self.report_milk = None
        

    def load_and_process(self):
        self.MA = get_dependency('milk_aggregates')
        self.WG = get_dependency('whiteboard_groups')
        self.report_milk = self.createReportMilk()

    def createReportMilk(self):
        tenday1 = self.MA.tenday
        halfday1 = self.MA.halfday
        groups1  = self.WG.whiteboard_groups_tenday

        # Drop the index
        tenday  = tenday1 .reset_index(drop=True)
        halfday = halfday1.reset_index(drop=True)
        groups  = groups1 .reset_index(drop=True)

        column_formats = {
            # Text columns (no formatting needed)
            'ultra': 'text',
            'group': 'text',

            # Float columns with 1 decimal place
            'average': '{:.1f}',
            'pct chg from avg': '{:.2f}',
            'chg from avg': '{:.2f}',
            'AM': '{:.1f}',
            'PM': '{:.1f}',
            'litres': '{:.1f}',
            'avg': '{:.1f}',

            # Float columns with no decimal places (integers)
            'WY_id': '{:.0f}',
            'cow_id': '{:.0f}',
            'milking days': '{:.0f}',
            'days milking': '{:.0f}',

            # date cols
            'expected bdate': 'iso8601',
            'bdate (exp)': 'iso8601',
        }

        def format_dataframe(df, formats):
            df_formatted = df.copy()

            if 'u_read' in df_formatted.columns:
                df_formatted['u_read'] = df_formatted['u_read'].fillna('')

            for col in df_formatted.columns:
                if col in formats:
                    if formats[col] == 'iso8601':
                        df_formatted[col] = pd.to_datetime(df_formatted[col], errors='coerce').dt.strftime('%Y-%m-%d')
                    elif formats[col] != 'text':
                        df_formatted[col] = pd.to_numeric(df_formatted[col], errors='coerce')
                        df_formatted[col] = df_formatted[col].apply(lambda x: formats[col].format(x) if pd.notna(x) else '')
                    else:
                        df_formatted[col] = df_formatted[col].astype(str)
                else:
                    df_formatted[col] = df_formatted[col].astype(str)
            return df_formatted

        tenday_formatted = format_dataframe(tenday, column_formats)
        halfday_formatted = format_dataframe(halfday, column_formats)
        groups_formatted = format_dataframe(groups, column_formats)

        # Save to Excel with separate sheets
        output_path = "F:\\COWS\\data\\milk_data\\groups\\report_milk.xlsx"

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            tenday_formatted.to_excel(writer, sheet_name='TenDay', index=False)
            halfday_formatted.to_excel(writer, sheet_name='HalfDay', index=False)
            groups_formatted.to_excel(writer, sheet_name='Groups', index=False)

        wb = load_workbook(output_path)
        for sheet_name in ['TenDay', 'HalfDay', 'Groups']:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='right')

            # Set column width for date columns
            for col in ws.columns:
                col_letter = col[0].column_letter
                header = col[0].value
                if header in ['expected bdate', 'bdate (exp)']:
                    ws.column_dimensions[col_letter].width = 15  # Set width to 15 (adjust as needed)

                    # Set header row height and wrap text
            header_row = ws[1]
            for cell in header_row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.row_dimensions[1].height = 50 

        wb.save(output_path)



        # Optionally, store for further use
        self.report_milk = {
            'TenDay': tenday_formatted,
            'HalfDay': halfday_formatted,
            'Groups': groups_formatted
        }
        return self.report_milk

    # gets written to output_path = "F:\\COWS\\data\\milk_data\\groups\\report_milk.xlsx"

if __name__ == "__main__":
    obj = ReportMilkXlsx()
    obj.load_and_process()    